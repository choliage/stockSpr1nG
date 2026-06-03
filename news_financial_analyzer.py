import json
import os
import re
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook

from keyword_scanner import KEYWORD_PATTERN, get_news_root

DEFAULT_BASE_DIR = os.path.join(os.path.expanduser("~"), "Desktop")


def _looks_like_company_code(value: str) -> bool:
    return bool(re.match(r"^[0-9]{3,6}[A-Z]?$", value.strip()))


def _load_company_names(company_names_path: Optional[str] = None) -> Dict[str, str]:
    if company_names_path is None:
        company_names_path = os.path.join(os.path.dirname(__file__), "company_names.json")

    if not os.path.exists(company_names_path):
        return {}

    with open(company_names_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        normalized = {str(k).strip(): str(v).strip() for k, v in data.items() if k}
        # 支援兩種格式：
        # 1) code -> name
        # 2) name -> code
        if normalized and all(
            not _looks_like_company_code(k) and _looks_like_company_code(v)
            for k, v in normalized.items()
        ):
            return {v: k for k, v in normalized.items()}
        return normalized

    if isinstance(data, list):
        return {str(item).strip(): "" for item in data if item}

    return {}


def _format_date(file_path: str) -> str:
    parent_dir = os.path.basename(os.path.dirname(file_path))
    if re.match(r"^\d{4}_\d{2}_\d{2}$", parent_dir):
        return parent_dir
    filename = os.path.splitext(os.path.basename(file_path))[0]
    if len(filename) >= 10 and re.match(r"^\d{4}_\d{2}_\d{2}$", filename[:10]):
        return filename[:10]
    try:
        return datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y_%m_%d")
    except Exception:
        return datetime.now().strftime("%Y_%m_%d")


def _collect_news_txt_files(base_dir: Optional[str] = None) -> List[str]:
    root = get_news_root(base_dir)
    if not os.path.isdir(root):
        return []
    txt_files = []
    for root_dir, _, files in os.walk(root):
        if os.path.basename(root_dir).lower() == "logs":
            continue
        for filename in files:
            if filename.lower().endswith(".txt"):
                txt_files.append(os.path.join(root_dir, filename))
    return txt_files


def _extract_keywords(text: str) -> List[str]:
    matches = {m.group(0) for m in KEYWORD_PATTERN.finditer(text)}
    return sorted(matches, key=lambda x: x.lower()) if matches else []


def _find_company_mentions(text: str, company_names: Dict[str, str]) -> List[Dict[str, str]]:
    mentions = []
    lower_text = text.lower()
    for code, name in company_names.items():
        found = False
        if code and code in text:
            mentions.append({"company_code": code, "company_name": name, "match_term": code})
            found = True
        if name:
            if name.lower() in lower_text:
                mentions.append({"company_code": code, "company_name": name, "match_term": name})
                found = True
        if found:
            continue
    return mentions


def collect_news_and_report_data(base_dir: Optional[str] = None, company_names_path: Optional[str] = None) -> Dict[str, object]:
    if base_dir is None:
        base_dir = DEFAULT_BASE_DIR

    company_names = _load_company_names(company_names_path)
    txt_files = _collect_news_txt_files(base_dir)

    company_data: Dict[str, Dict[str, object]] = {}
    overall_rows = []
    overall_keyword_counts: Dict[str, int] = {}

    for txt_file in txt_files:
        try:
            with open(txt_file, "r", encoding="utf-8") as f:
                text = f.read()
        except Exception:
            continue

        date_str = _format_date(txt_file)
        keywords = _extract_keywords(text)
        for kw in keywords:
            overall_keyword_counts[kw] = overall_keyword_counts.get(kw, 0) + 1

        mentions = _find_company_mentions(text, company_names)
        if not mentions:
            mentions = [{"company_code": "UNKNOWN", "company_name": "", "match_term": ""}]

        for mention in mentions:
            code = mention["company_code"]
            name = mention.get("company_name", "")
            row = {
                "date": date_str,
                "file_name": os.path.basename(txt_file),
                "company_code": code,
                "company_name": name,
                "match_term": mention.get("match_term", ""),
                "keywords": ", ".join(keywords),
                "keyword_count": len(keywords),
            }
            overall_rows.append(row)

            if code not in company_data:
                company_data[code] = {
                    "company_name": name,
                    "mention_count": 0,
                    "keyword_count": 0,
                    "files": set(),
                    "keyword_set": set(),
                }
            company_data[code]["mention_count"] += 1
            company_data[code]["keyword_count"] += len(keywords)
            company_data[code]["files"].add(row["file_name"])
            company_data[code]["keyword_set"].update(keywords)

    report_root = os.path.join(base_dir, "Financial_Reports")
    report_summary: Dict[str, Dict[str, object]] = {}
    if os.path.isdir(report_root):
        for code_dir in os.listdir(report_root):
            code_path = os.path.join(report_root, code_dir)
            if not os.path.isdir(code_path):
                continue
            report_files = [f for f in os.listdir(code_path) if f.lower().endswith(".pdf")]
            if not report_files:
                continue
            report_summary[code_dir] = {
                "report_count": len(report_files),
                "latest_report": max(report_files),
            }
            if code_dir not in company_data:
                company_data[code_dir] = {
                    "company_name": company_names.get(code_dir, ""),
                    "mention_count": 0,
                    "keyword_count": 0,
                    "files": set(),
                    "keyword_set": set(),
                }

    combined_rows = []
    for code, data in sorted(company_data.items(), key=lambda item: item[0]):
        rpt = report_summary.get(code, {})
        combined_rows.append(
            {
                "company_code": code,
                "company_name": data.get("company_name", ""),
                "mention_count": data.get("mention_count", 0),
                "keyword_count": data.get("keyword_count", 0),
                "unique_news_files": len(data.get("files", set())),
                "keywords": ", ".join(sorted(data.get("keyword_set", set()))),
                "report_count": rpt.get("report_count", 0),
                "latest_report": rpt.get("latest_report", ""),
            }
        )

    return {
        "company_rows": combined_rows,
        "article_rows": overall_rows,
        "overall_keyword_counts": overall_keyword_counts,
        "company_names_path": os.path.abspath(company_names_path) if company_names_path else "",
    }


def save_analysis_excel(base_dir: Optional[str] = None, company_names_path: Optional[str] = None) -> str:
    if base_dir is None:
        base_dir = DEFAULT_BASE_DIR

    analysis_root = os.path.join(base_dir, "News_Financial_Analysis")
    os.makedirs(analysis_root, exist_ok=True)

    data = collect_news_and_report_data(base_dir=base_dir, company_names_path=company_names_path)
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Company Summary"
    summary_ws.append([
        "公司代號",
        "公司名稱",
        "新聞提及次數",
        "關鍵字匹配次數",
        "不同新聞檔案數",
        "關鍵字清單",
        "已下載財報數",
        "最新財報檔名",
    ])

    for row in data["company_rows"]:
        summary_ws.append([
            row["company_code"],
            row["company_name"],
            row["mention_count"],
            row["keyword_count"],
            row["unique_news_files"],
            row["keywords"],
            row["report_count"],
            row["latest_report"],
        ])

    detail_ws = workbook.create_sheet("News Detail")
    detail_ws.append([
        "日期",
        "新聞 txt 檔案",
        "公司代號",
        "公司名稱",
        "匹配字詞",
        "關鍵字",
        "關鍵字數量",
    ])
    for row in data["article_rows"]:
        detail_ws.append([
            row["date"],
            row["file_name"],
            row["company_code"],
            row["company_name"],
            row["match_term"],
            row["keywords"],
            row["keyword_count"],
        ])

    counts_ws = workbook.create_sheet("Keyword Counts")
    counts_ws.append(["關鍵字", "出現次數"])
    for keyword, count in sorted(data["overall_keyword_counts"].items(), key=lambda item: (-item[1], item[0])):
        counts_ws.append([keyword, count])

    output_path = os.path.join(analysis_root, "news_financial_analysis.xlsx")
    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    path = save_analysis_excel()
    print(f"已產生分析檔案：{path}")
